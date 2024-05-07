import openpyxl

wb = openpyxl.load_workbook('./12excel/produceSales.xlsx')
sheet = wb['Sheet']

PRICE_UPDATE = {'Garlic': 3.07,
                'Celery': 1.19,
                'Lemon': 1.27}

for row_num in range(2, sheet.max_row + 1): # 先頭行をスキップ
    produce_name = sheet.cell(row=row_num, column=1).value
    if produce_name in PRICE_UPDATE:
        sheet.cell(row=row_num, column=2).value = PRICE_UPDATE[produce_name]
        
wb.save('updateProduceSales.xlsx')